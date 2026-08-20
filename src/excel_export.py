"""銘柄コードから「企業詳細」「実行表」のExcelを自動生成するモジュール。

J-Quantsから取得できる数値項目（時価総額・PBR・PER・配当利回り・株価・
四半期業績等）だけを自動入力する。事業概要・主要株主・将来予想・定性コメント
等の自由記述はJ-Quantsだけでは自動化できないため、生成後にユーザーが手動で
追記する運用とする（README.md参照）。
"""
from __future__ import annotations

import datetime as dt
import io
import logging
import math
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.comments import Comment

from src import edinet_client, endpoints, pipeline, regional_stocks, tdnet_client
from src.config import REGIONAL_LISTING_LOOKBACK_YEARS
from src.jquants_client import JQuantsClient

logger = logging.getLogger(__name__)

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"
COMPANY_DETAIL_TEMPLATE = TEMPLATE_DIR / "company_detail_template.xlsx"

# equities/masterのProdCat（商品区分）。011=内国株式（普通株式）で、それ以外は
# ETF(014)・REIT(013)・ETN(023)・外国株式(021)・日本銀行等(012)等、事業内容や
# 決算を持たない/対象外の銘柄のため、企業詳細・実行表の生成対象から除外する。
_COMMON_STOCK_PROD_CAT = "011"


class NotCommonStockError(Exception):
    """銘柄コードがETF・REIT・ETN等、株式会社の普通株式ではない場合。"""


def _to_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _get_master_row(client: JQuantsClient, code: str) -> dict:
    listed = endpoints.get_listed_info(client)
    if listed.empty or "Code" not in listed.columns:
        return {}
    # equities/masterのCodeは5桁表記（4桁コードは末尾に"0"が付く）だが、
    # fins/summary・equities/bars/dailyは4桁のままでも通るため、入力側は
    # 4桁のことがある。末尾"0"を補った表記も候補にして突き合わせる。
    code_str = str(code)
    candidates = {code_str}
    if code_str.isdigit() and len(code_str) == 4:
        candidates.add(code_str + "0")
    row = listed.loc[listed["Code"].astype(str).isin(candidates)]
    return row.iloc[0].to_dict() if not row.empty else {}


def _regional_fallback_info(code: str) -> tuple[str, str | None]:
    """equities/masterに載っていない銘柄（地方取引所単独上場等）向けに、
    TDnet開示から会社名と直近の上場市場情報(markets_string)を補う。

    地方単独上場企業の検出に使っているものと同じTDnetデータ（キャッシュ済み
    分は再利用、無ければ取得）を使う。見つからない場合は("", None)を返す
    （誤った推測値を出さない。README「地方株」セクション参照）。
    """
    end = dt.date.today()
    start = end - dt.timedelta(days=365 * REGIONAL_LISTING_LOOKBACK_YEARS)
    try:
        # 当日分のTDnet開示はまだ全件公開されていない可能性がある。
        # tdnet_clientは期間指定をそのままキャッシュキーにするため、当日を
        # 含む期間を一度キャッシュしてしまうと同じ銘柄で何度Excelを生成しても
        # その日のうちは同じ不完全な結果を返し続けてしまう。前日までとは別に、
        # 当日分だけforce_refresh=Trueで毎回取り直す（src/regional_stocks.pyの
        # update_regional_storeと同じパターン。2026-08-20のCodexレビューで指摘）。
        yesterday = end - dt.timedelta(days=1)
        stable = tdnet_client.get_disclosures_range(start, yesterday) if start <= yesterday else pd.DataFrame()
        todays = tdnet_client.get_disclosures_range(end, end, force_refresh=True)
        disclosures = pd.concat([stable, todays], ignore_index=True)
    except Exception as exc:  # noqa: BLE001 -- TDnetミラー障害時も他の処理は続行する
        logger.info("[%s] TDnetからの会社名/市場情報の取得に失敗しました: %s", code, exc)
        return "", None
    if disclosures.empty or "company_code" not in disclosures.columns:
        return "", None
    # TDnetのcompany_codeもequities/masterと同じ5桁表記（4桁コードは英数字を
    # 問わず末尾に"0"が付く。例: "231A"→"231A0"）で来ることがあるため、
    # _get_master_rowと同じ候補突き合わせをする。
    code_str = str(code)
    candidates = {code_str}
    if len(code_str) == 4:
        candidates.add(code_str + "0")
    rows = disclosures.loc[disclosures["company_code"].astype(str).isin(candidates)]
    if rows.empty:
        return "", None
    rows = rows.copy()
    rows["pubdate"] = pd.to_datetime(rows["pubdate"], errors="coerce")
    rows = rows.sort_values("pubdate")
    name = rows.iloc[-1].get("company_name") or ""
    # markets_stringが欠損している回（会社名だけの開示等）が直近の開示だと、
    # それより前の回で分かっている市場情報を取り逃してしまう
    # （src/regional_stocks.py の_latest_company_statusと同じ既知の欠損
    # パターン。2026-08-20のCodexレビューで指摘）。会社名は単純に最新の開示
    # から、markets_stringは有効な値を持つ最新の開示から、それぞれ独立に選ぶ。
    has_valid_market = rows["markets_string"].apply(lambda m: isinstance(m, str) and bool(m))
    market_rows = rows.loc[has_valid_market]
    markets_string = market_rows.iloc[-1]["markets_string"] if not market_rows.empty else None
    return str(name), (str(markets_string) if markets_string is not None else None)


def _regional_fallback_price(code: str, markets_string: str | None) -> tuple[float | None, dt.date | None]:
    """地方単独上場企業向けのyfinance株価フォールバック（現在値のみ、実機確認済み
    な福証銘柄のみ成功しうる。src/regional_stocks.pyのfetch_regional_share_priceと
    同じロジックを再利用する）。取得できない場合は(None, None)を返す。
    """
    if not markets_string:
        return None, None
    price, _note = regional_stocks.fetch_regional_share_price(code, markets_string)
    if price is None:
        return None, None
    return price, dt.date.today()


def _ensure_common_stock(master_row: dict, code: str) -> None:
    prod_cat = master_row.get("ProdCat")
    if master_row and prod_cat != _COMMON_STOCK_PROD_CAT:
        name = master_row.get("CoName", "")
        raise NotCommonStockError(
            f"銘柄コード{code}（{name}）はETF・REIT・ETN等の金融商品であり、株式会社の"
            "普通株式ではありません。企業詳細・実行表は株式会社向けの項目（事業概要・決算等）"
            "のみに対応しています。"
        )


def _nearest_close(prices: pd.DataFrame, target_date) -> float | None:
    """target_date以前で最も近い（無ければ最初の）取引日の終値を返す。"""
    if prices.empty or target_date is None or pd.isna(target_date):
        return None
    target = pd.Timestamp(target_date)
    on_or_before = prices.loc[prices["Date"] <= target]
    row = on_or_before.iloc[-1] if not on_or_before.empty else prices.iloc[0]
    close = _to_numeric(pd.Series([row.get("C")])).iloc[0]
    return None if pd.isna(close) else float(close)


def _stock_split_events_text(prices: pd.DataFrame) -> str | None:
    """株価履歴のAdjFactor（調整係数）から株式分割・併合があった日付を検出する
    （rules.detect_stock_splitと同じ考え方。契約プランの取得可能期間より前の
    分割・併合は株価データが無く検出できない）。
    """
    if prices.empty or "AdjFactor" not in prices.columns or "Date" not in prices.columns:
        return None
    df = prices.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["AdjFactor"] = _to_numeric(df["AdjFactor"])
    df = df.dropna(subset=["Date", "AdjFactor"]).sort_values("Date")
    hits = df.loc[df["AdjFactor"] != 1.0]
    if hits.empty:
        return None
    lines = [
        f"{row['Date'].date()}（調整係数 {row['AdjFactor']}）"
        for _, row in hits.iterrows()
    ]
    return "株式分割・併合: " + "、".join(lines)



def _latest_actual_fy_end(fins: pd.DataFrame) -> dt.date | None:
    """本決算(FY)の実績開示（Salesが入っている行）に基づく直近の決算期末日を返す。

    配当予想の修正等、Salesを伴わない開示にはCurFYEnが未来の（まだ実績が
    確定していない）決算期を指すものがあり、それを「最新の決算期」と
    誤認してEDINET検索の窓がずれる（未来日付になり書類が見つからない）
    ことがあるため、実績Salesが入っている本決算行だけから求める。
    """
    if fins.empty or not {"CurFYEn", "CurPerType", "Sales"}.issubset(fins.columns):
        return None
    f = fins.copy()
    f["CurFYEn"] = pd.to_datetime(f["CurFYEn"], errors="coerce")
    f["Sales"] = _to_numeric(f["Sales"])
    f = f.dropna(subset=["CurFYEn", "Sales"])
    f = f.loc[f["CurPerType"] == "FY"]
    if f.empty:
        return None
    return f["CurFYEn"].max().date()


_YFINANCE_ESTIMATE_NOTE = (
    "この値はJ-Quantsに株価データが無い銘柄（地方取引所単独上場等）向けに、"
    "yfinance（非公式データ源）の現在値を使って計算した参考値です。"
    "J-Quantsで検証された実績値ではありません。"
)


def _add_yfinance_note(ws, *cell_refs: str) -> None:
    for ref in cell_refs:
        ws[ref].comment = Comment(_YFINANCE_ESTIMATE_NOTE, "excel_export.py")


def build_company_detail_excel(client: JQuantsClient, code: str) -> bytes:
    """銘柄コードから「企業詳細」シートを埋めたExcel(bytes)を生成する。"""
    code = str(code)
    master_row = _get_master_row(client, code)
    _ensure_common_stock(master_row, code)
    fins = endpoints.get_financials_by_code(client, code)
    prices = endpoints.get_price_history_by_code(client, code)

    company_name = master_row.get("CoName", "")
    regional_markets_string = None
    if not master_row:
        # equities/masterに無い銘柄（地方取引所単独上場等）はTDnetから会社名・
        # 市場情報を補う（README「地方株」セクション参照）。
        company_name, regional_markets_string = _regional_fallback_info(code)

    fallback_price = fallback_price_date = None
    if prices.empty and regional_markets_string:
        fallback_price, fallback_price_date = _regional_fallback_price(code, regional_markets_string)

    metrics = pipeline.compute_market_metrics(fins, prices, fallback_price, fallback_price_date)
    listing_date, _recently_listed = pipeline.estimate_listing_date(prices)
    used_yfinance = metrics.get("price_source") == "yfinance"

    wb = openpyxl.load_workbook(COMPANY_DETAIL_TEMPLATE)
    ws = wb["原本"]

    ws["B1"] = code
    ws["E1"] = company_name

    latest_fy_end = _latest_actual_fy_end(fins)
    if latest_fy_end is not None:
        ws["O1"] = f"{latest_fy_end.month}月"

    market_cap = metrics["market_cap"]
    if market_cap is not None and pd.notna(market_cap):
        ws["E2"] = round(market_cap / 1e8, 1)
    pbr = metrics["pbr"]
    if pbr is not None and pd.notna(pbr):
        ws["I2"] = round(pbr, 2)
    per = metrics["per"]
    if per is not None and pd.notna(per):
        ws["M2"] = round(per, 2)
    dividend_yield = metrics["dividend_yield"]
    if dividend_yield is not None and pd.notna(dividend_yield):
        # 表示形式(0.0%)による四捨五入を避けるため、小数点第1位（0.1%単位）に
        # あらかじめ切り捨てておく。
        ws["Q2"] = math.trunc(dividend_yield * 1000) / 1000

    latest_close = metrics["latest_close"]
    if latest_close is not None and pd.notna(latest_close):
        ws["B3"] = round(latest_close)

    if used_yfinance:
        _add_yfinance_note(ws, "E2", "I2", "M2", "Q2", "B3")

    if listing_date is not None:
        ws["F3"] = listing_date
        ws["F3"].number_format = "yyyy/m/d"

    shares_out = metrics["shares_out"]
    if shares_out is not None and pd.notna(shares_out):
        ws["H4"] = round(shares_out / 1e4)

    market_name = master_row.get("MktNm")
    if market_name:
        ws["O4"] = market_name

    yuho = edinet_client.fetch_yuho_texts(code, latest_fy_end)

    if yuho["business_overview"]:
        ws["C7"] = yuho["business_overview"]
    if yuho["shareholders"] is not None and not yuho["shareholders"].empty:
        ws["C9"] = _format_shareholders_text(yuho["shareholders"])

    split_text = _stock_split_events_text(prices)
    c12_parts = [t for t in [yuho["potential_shares"], split_text] if t]
    if c12_parts:
        ws["C12"] = "\n\n".join(c12_parts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _format_shareholders_text(shareholders: pd.DataFrame, top_n: int = 6) -> str:
    """大株主の状況テーブル(EDINET有報)を上位N件の読みやすい複数行テキストに整形する。"""
    lines = []
    for _, row in shareholders.head(top_n).iterrows():
        parts = [str(v) for v in row.tolist() if pd.notna(v) and str(v).strip()]
        lines.append("　".join(parts))
    return "\n".join(lines)


EXECUTION_TABLE_TEMPLATE = TEMPLATE_DIR / "execution_table_template.xlsx"

# テンプレート(templates/execution_table_template.xlsx)の固定レイアウト。
# 年度列は7枠(C,E,G,I,K,M,O)あり、C列から左詰めで実績年度を並べ、会社予想
# 年度（今期・来期の2年分まで、開示されている範囲で）はその直後の列に置く
# （Oの右隣に単位列があるのと同じ構成）。
_YEAR_COLS = ["C", "E", "G", "I", "K", "M", "O"]
_ROW_SALES = {"1Q": 5, "2Q": 6, "3Q": 7, "FY": 8}
_ROW_PROFIT = {"1Q": 9, "2Q": 10, "3Q": 11, "FY": 12}
# テンプレートの利益セルは元々小数1桁の書式("0.0;[Red]▲0.0;-")のため、
# 小数2桁で丸めた値(round(profit / 1e8, 2))をそのまま書き込んでもExcel上は
# 1桁までしか表示されない。値と書式を同時に更新する必要がある
# （2026-08-20のCodexレビューで指摘）。
_PROFIT_NUMBER_FORMAT = "0.00;[Red]▲0.00;-"
_ROW_PRICE = {"1Q": 14, "2Q": 15, "3Q": 16, "FY": 17}


def build_execution_table_excel(client: JQuantsClient, code: str) -> bytes:
    """銘柄コードから「実行表」（年度×四半期の売上・経常利益・株価マトリクス）
    を埋めたExcel(bytes)を生成する。

    テンプレート(templates/execution_table_template.xlsx、元の三桜工業実行表から
    個社データを除いたもの)のレイアウト・書式・数式(利益率・検算)をそのまま使い、
    値セルだけを埋める。対象年度は契約プラン（Lightは過去5年）でJ-Quantsから
    取得できる実績年度（直近5年分）＋今期・来期の予想年度2年分に自動で限定
    される（テンプレートの列数7枠が上限）。予想の2列は会社予想が未開示でも
    見出しだけ用意し、値は開示されていれば自動で、無ければ手動で入力する。
    それより古い実績年度は手動での追記が必要。
    """
    code = str(code)
    master_row = _get_master_row(client, code)
    _ensure_common_stock(master_row, code)
    fins = endpoints.get_financials_by_code(client, code)
    prices_raw = endpoints.get_price_history_by_code(client, code)

    company_name = master_row.get("CoName", "")
    regional_markets_string = None
    if not master_row:
        # equities/masterに無い銘柄（地方取引所単独上場等）はTDnetから会社名・
        # 市場情報を補う（README「地方株」セクション参照）。
        company_name, regional_markets_string = _regional_fallback_info(code)

    fallback_price = fallback_price_date = None
    if prices_raw.empty and regional_markets_string:
        fallback_price, fallback_price_date = _regional_fallback_price(code, regional_markets_string)

    metrics = pipeline.compute_market_metrics(fins, prices_raw, fallback_price, fallback_price_date)
    used_yfinance = metrics.get("price_source") == "yfinance"

    prices = prices_raw.copy()
    if not prices.empty:
        prices["Date"] = pd.to_datetime(prices["Date"], errors="coerce")
        prices = prices.dropna(subset=["Date"]).sort_values("Date")

    wb = openpyxl.load_workbook(EXECUTION_TABLE_TEMPLATE)
    ws = wb["1"]

    ws["B1"] = code
    ws["C1"] = company_name
    latest_fy_end = _latest_actual_fy_end(fins)
    if latest_fy_end is not None:
        ws["M1"] = latest_fy_end.month

    if metrics["market_cap"] is not None and pd.notna(metrics["market_cap"]):
        ws["B2"] = round(metrics["market_cap"] / 1e8, 1)
    if metrics["pbr"] is not None and pd.notna(metrics["pbr"]):
        ws["F2"] = round(metrics["pbr"], 2)
    if metrics["per"] is not None and pd.notna(metrics["per"]):
        ws["J2"] = round(metrics["per"], 2)
    if metrics["dividend_yield"] is not None and pd.notna(metrics["dividend_yield"]):
        # 表示形式(0.0)による四捨五入を避けるため、小数点第1位にあらかじめ
        # 切り捨てておく。
        ws["O2"] = math.trunc(metrics["dividend_yield"] * 1000) / 10
    if metrics["latest_close"] is not None and pd.notna(metrics["latest_close"]):
        ws["B3"] = round(metrics["latest_close"])

    if used_yfinance:
        _add_yfinance_note(ws, "B2", "F2", "J2", "O2", "B3")

    required = {"CurFYEn", "CurPerType", "DiscDate", "Sales", "OdP"}
    if fins.empty or not required.issubset(fins.columns):
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    f = fins.copy()
    f["CurFYEn"] = pd.to_datetime(f["CurFYEn"], errors="coerce")
    f["CurPerEn"] = pd.to_datetime(f.get("CurPerEn"), errors="coerce")
    f["DiscDate"] = pd.to_datetime(f["DiscDate"], errors="coerce")
    f["Sales"] = _to_numeric(f["Sales"])
    f["OdP"] = _to_numeric(f["OdP"])
    f["OP"] = _to_numeric(f["OP"]) if "OP" in f.columns else float("nan")
    f = f.dropna(subset=["CurFYEn"])

    # 決算短信以外の開示（配当予想の修正等）はSales/OdPが空でもCurFYEnを持つため、
    # 実績データの年度は「Salesが実際に入っている行」だけから拾う。
    f_actual = f.dropna(subset=["Sales"])
    # 「実績年度」（見出しに"予想"を付けず、今期・来期予想欄の基準にもする年度）
    # は、本決算(FY)自体の実績が確定した年度だけに限る。決算期が3月以外の会社等
    # では、進行中の期でも1Q/2Q/3Qの実績Salesは入っているため、これを含めると
    # 本決算がまだ来ていない進行中の期まで「実績」と誤認し、実際には会社予想
    # （FSales/FOdP）で埋まっているだけの決算行が実績のように見えたり、今期・
    # 来期の予想欄が1期ズレて空欄になったりする。
    f_actual_fy = f_actual.loc[f_actual["CurPerType"] == "FY"]
    all_actual_years = sorted(f_actual_fy["CurFYEn"].dt.year.unique().tolist())

    # 「今期」の会社予想(FSales/FOdP)・「来期」の会社予想(NxFSales/NxFOdP)は、
    # CurPerType="FY"の開示（本決算実績、または本決算の会社予想を修正する
    # EarnForecastRevision等）でのみ年間換算の値になる。四半期(1Q/2Q/3Q)を
    # 対象にしたEarnForecastRevisionではFSales等がその四半期だけの予想値に
    # なることがあるため、それらは使わずCurPerType="FY"の開示だけから拾う
    # （本決算実績の開示時点ではFSales/FOdPは実績確定済みのため空になり、
    # NxFSales/NxFOdPだけが次期予想として入っている）。
    forecast_candidates: list[tuple[int, float, float, pd.Timestamp]] = []
    fy_rows = f.loc[f["CurPerType"] == "FY"]
    if not fy_rows.empty:
        latest_fy_row = fy_rows.sort_values("DiscDate").iloc[-1]
        disc_date = latest_fy_row["DiscDate"]

        cur_fy_end = latest_fy_row["CurFYEn"]
        fsales = _to_numeric(pd.Series([latest_fy_row.get("FSales")])).iloc[0]
        fodp = _to_numeric(pd.Series([latest_fy_row.get("FOdP")])).iloc[0]
        if pd.isna(fodp):
            # IFRS採用企業には経常利益予想の区分が無いため、営業利益予想(FOP)で代用する。
            fodp = _to_numeric(pd.Series([latest_fy_row.get("FOP")])).iloc[0]
        if pd.notna(cur_fy_end) and pd.notna(fsales):
            forecast_candidates.append((int(cur_fy_end.year), fsales, fodp, disc_date))

        nxt_fy_end = pd.to_datetime(latest_fy_row.get("NxtFYEn"), errors="coerce")
        nx_fsales = _to_numeric(pd.Series([latest_fy_row.get("NxFSales")])).iloc[0]
        nx_fodp = _to_numeric(pd.Series([latest_fy_row.get("NxFOdP")])).iloc[0]
        if pd.isna(nx_fodp):
            # IFRS採用企業には経常利益予想の区分が無いため、営業利益予想(NxFOP)で代用する。
            nx_fodp = _to_numeric(pd.Series([latest_fy_row.get("NxFOP")])).iloc[0]
        if pd.notna(nxt_fy_end) and pd.notna(nx_fsales):
            forecast_candidates.append((int(nxt_fy_end.year), nx_fsales, nx_fodp, disc_date))

    forecast_data: dict[int, tuple[float, float, pd.Timestamp]] = {
        year: (sales, profit, disc_date)
        for year, sales, profit, disc_date in forecast_candidates
        if year not in all_actual_years
    }
    logger.info(
        "[%s] 最新実績決算期=%s 会社予想決算期=%s 会社予想売上=%s 会社予想経常利益=%s",
        code,
        all_actual_years[-1] if all_actual_years else None,
        sorted(forecast_data.keys()),
        {y: v[0] for y, v in forecast_data.items()},
        {y: v[1] for y, v in forecast_data.items()},
    )

    # 今期・来期（直近の実績年度の次の2年）は、会社予想がまだ開示されていない
    # 場合でも「YYYY年予想」の列見出し自体は常に用意する（値は空欄のまま手動
    # 入力できるようにする）。開示済みの予想がある場合はforecast_dataの値で
    # 自動的に埋まる。
    last_actual_year = max(all_actual_years) if all_actual_years else None
    default_forecast_years = (
        [last_actual_year + 1, last_actual_year + 2] if last_actual_year is not None else []
    )
    forecast_years = sorted(
        {
            year
            for year in set(forecast_data.keys()) | set(default_forecast_years)
            if year not in all_actual_years
        }
    )[:2]

    if not all_actual_years and not forecast_years:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # テンプレートの年度枠は7つ。予想年度分の枠(今期・来期の2つ)を確保した
    # 残りは、直近の実績年度を起点にした「連続したカレンダー年」を固定で
    # C列から並べる（実績が無い年でも列自体は必ず用意し、その年だけ空欄に
    # する）。データが存在する年だけを詰めて並べると、欠けている年がある
    # 銘柄（市場変更・上場間もない等）で列と年の対応がずれてしまうため。
    max_actual = max(len(_YEAR_COLS) - len(forecast_years), 0)
    actual_years = (
        list(range(last_actual_year - max_actual + 1, last_actual_year + 1))
        if last_actual_year is not None and max_actual
        else []
    )

    col_for_year = {}
    for i, year in enumerate(actual_years):
        col_for_year[year] = _YEAR_COLS[i]
    for i, year in enumerate(forecast_years):
        idx = len(actual_years) + i
        if idx < len(_YEAR_COLS):
            col_for_year[year] = _YEAR_COLS[idx]

    for year, col in col_for_year.items():
        ws[f"{col}4"] = f"{year}年予想" if year in forecast_years else f"{year}年"

    for _, row in f_actual.sort_values(["CurFYEn", "CurPerType"]).iterrows():
        period_type = row["CurPerType"]
        if period_type not in _ROW_SALES:
            continue
        year = int(row["CurFYEn"].year)
        col = col_for_year.get(year)
        if col is None:
            continue
        sales = row["Sales"]
        profit = row["OdP"]
        if pd.isna(profit) and pd.notna(row.get("OP")):
            # IFRS採用企業には経常利益の区分が無いため、営業利益(OP)で代用する。
            profit = row["OP"]
        if pd.notna(sales):
            sales_cell = ws[f"{col}{_ROW_SALES[period_type]}"]
            sales_cell.value = math.floor(sales / 1e8)
            sales_cell.number_format = "0;[Red]▲0;-"
        if pd.notna(profit):
            profit_cell = ws[f"{col}{_ROW_PROFIT[period_type]}"]
            profit_cell.value = round(profit / 1e8, 2)
            profit_cell.number_format = _PROFIT_NUMBER_FORMAT
        close = _nearest_close(prices, row["CurPerEn"])
        if close is not None:
            ws[f"{col}{_ROW_PRICE[period_type]}"] = round(close)

    for year, (sales, profit, disc_date) in forecast_data.items():
        col = col_for_year.get(year)
        if col is None:
            continue
        if pd.notna(sales):
            forecast_sales_cell = ws[f"{col}{_ROW_SALES['FY']}"]
            forecast_sales_cell.value = math.floor(sales / 1e8)
            forecast_sales_cell.number_format = "0;[Red]▲0;-"
        if pd.notna(profit):
            forecast_profit_cell = ws[f"{col}{_ROW_PROFIT['FY']}"]
            forecast_profit_cell.value = round(profit / 1e8, 2)
            forecast_profit_cell.number_format = _PROFIT_NUMBER_FORMAT

    # 決算期が3月以外の会社等では、直近の実績年度がまだ本決算を迎えておらず
    # （1Q/2Q等の実績しかない）「決算」行が空欄になることがある。その場合は
    # 直近の開示に入っている会社予想(FSales/FOdP、その期自体の予想)で埋める
    # （本決算が実際に出れば、次回生成時に実績値へ自動的に置き換わる）。
    for year in actual_years:
        col = col_for_year.get(year)
        if col is None:
            continue
        if ws[f"{col}{_ROW_SALES['FY']}"].value is not None:
            continue
        year_rows = f_actual.loc[f_actual["CurFYEn"].dt.year == year]
        if year_rows.empty:
            continue
        latest_row = year_rows.sort_values("DiscDate").iloc[-1]
        fsales = _to_numeric(pd.Series([latest_row.get("FSales")])).iloc[0]
        fodp = _to_numeric(pd.Series([latest_row.get("FOdP")])).iloc[0]
        if pd.isna(fodp):
            # IFRS採用企業には経常利益予想の区分が無いため、営業利益予想(FOP)で代用する。
            fodp = _to_numeric(pd.Series([latest_row.get("FOP")])).iloc[0]
        if pd.notna(fsales):
            cell = ws[f"{col}{_ROW_SALES['FY']}"]
            cell.value = math.floor(fsales / 1e8)
            cell.number_format = "0;[Red]▲0;-"
        if pd.notna(fodp):
            fy_profit_cell = ws[f"{col}{_ROW_PROFIT['FY']}"]
            fy_profit_cell.value = round(fodp / 1e8, 2)
            fy_profit_cell.number_format = _PROFIT_NUMBER_FORMAT

    logger.info(
        "[%s] 会社予想EPS=%s 会社予想年間配当=%s 最新株価=%s PER=%s 配当利回り=%s 列割当=%s",
        code,
        pipeline._last_valid_full_year_value(f, "FEPS"),
        pipeline._dividend_forecast_or_trailing(f),
        metrics["latest_close"],
        metrics["per"],
        metrics["dividend_yield"],
        col_for_year,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
